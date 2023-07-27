from rest_framework import viewsets, permissions, pagination, generics
from rest_framework.decorators import api_view
from rest_framework.views import APIView
from rest_framework.response import Response
from openpyxl import load_workbook
from .serializers import (
    QuestionCategorySerializer,
    QuestionSerializer,
    QuestionOptionSerializer,
    ExcelUploadSerializer,
    QuizCategorySerializer,
    QuizSerializer,
    QuizQuestionSerializer,
    QuizQuestion,
    QuizQuestionGroupSerializer,
    BulkQuizQuestionSerializer,
    AllowedUserSerializer,
    UserQuizSerializer,
    UserAttemptSerializer,
)
import base64
from django.core.files.base import ContentFile
from django.shortcuts import get_object_or_404
from django.http import Http404
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Max, Q
from quiz.api.crud import create_user_attempt
from quiz.models import (
    QuestionCategory,
    Question,
    QuestionOption,
    QuizCategory,
    Quiz,
    QuizQuestion,
    QuizQuestionGroup,
    AllowedGrade,
    AllowedUser,
    UserAttempt,
    QuizInstanceQuestion,
    QuizInstanceOption,
)

from rest_framework.filters import SearchFilter
from .filters import CustomIdFilter, CustomQuestionFilter
from .permissions import StartQuizPermission
import random
from django.contrib.auth import get_user_model

User = get_user_model()


# Question views
def format_base64(request):
    image_data: str = request.data.get("body_photo")

    if image_data:
        image_binary = base64.b64decode(image_data.split(",")[-1])

        image_file = ContentFile(image_binary, name="filename.jpg")

        request.data["body_photo"] = image_file
    return request


class Base64PhotoMixin:
    def create(self, request, *args, **kwargs):
        request = format_base64(request)

        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        request = format_base64(request)

        return super().update(request, *args, **kwargs)


class QuestionsPagination(pagination.PageNumberPagination):
    page_size = 20
    page_size_query_param = "per_page"
    max_page_size = 10000


class QuestionCategoryViewSet(viewsets.ModelViewSet):
    queryset = QuestionCategory.objects.order_by("-updated")
    serializer_class = QuestionCategorySerializer
    permission_classes = [permissions.IsAdminUser]


class QuestionViewSet(Base64PhotoMixin, viewsets.ModelViewSet):
    queryset = Question.objects.order_by("-updated")
    serializer_class = QuestionSerializer
    pagination_class = QuestionsPagination

    filter_backends = (
        CustomQuestionFilter,
        CustomIdFilter,
        SearchFilter,
    )
    search_fields = ["id", "body_text", "options__body_text"]


class QuestionOptionViewSet(Base64PhotoMixin, viewsets.ModelViewSet):
    queryset = QuestionOption.objects.order_by("order_number")
    serializer_class = QuestionOptionSerializer
    filterset_fields = ["question"]


class ExcelUploadView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def post(self, request):
        serializer = ExcelUploadSerializer(data=request.data)
        if serializer.is_valid():
            excel_file = serializer.validated_data["excel_file"]
            category_id = serializer.validated_data["category"]
            options = serializer.validated_data["options"]

            wb = load_workbook(excel_file)
            ws = wb.active
            orders = [i for i in range(1, options + 1)]

            for row in ws.iter_rows(min_row=2, values_only=True):
                random.shuffle(orders)
                if category_id != 0:
                    new_question = Question.objects.create(
                        category_id=category_id, body_text=row[0]
                    )
                else:
                    new_question = Question.objects.create(body_text=row[0])
                QuestionOption.objects.create(
                    question=new_question,
                    body_text=row[1],
                    order_number=orders[0],
                    is_correct=True,
                )
                for i in range(2, options + 1):
                    QuestionOption.objects.create(
                        question=new_question,
                        body_text=row[i],
                        order_number=orders[i - 1],
                    )
            return Response({"status": "success"}, status=201)
        else:
            return Response({"status": serializer.errors}, status=400)


# Quiz views


class QuizCategoryViewset(viewsets.ModelViewSet):
    queryset = QuizCategory.objects.order_by("-updated")
    serializer_class = QuizCategorySerializer
    permission_classes = [permissions.IsAdminUser]


class QuizPagination(pagination.PageNumberPagination):
    page_size = 20
    page_size_query_param = "per_page"
    max_page_size = 10000


class QuizViewSet(viewsets.ModelViewSet):
    queryset = Quiz.objects.order_by("start_time")
    serializer_class = QuizSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ["category"]
    search_fields = ["title", "description"]
    permission_classes = [permissions.IsAdminUser]
    pagination_class = QuizPagination


class QuizQuestionPagination(pagination.PageNumberPagination):
    page_size = 20
    page_size_query_param = "per_page"
    max_page_size = 10000


class QuizQuestionViewset(viewsets.ModelViewSet):
    queryset = QuizQuestion.objects.order_by("order_number")
    serializer_class = QuizQuestionSerializer
    permission_classes = [permissions.IsAdminUser]
    pagination_class = QuizPagination
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["quiz"]


class QuizQuestionGroupViewSet(viewsets.ModelViewSet):
    queryset = QuizQuestionGroup.objects.order_by("order_number")
    serializer_class = QuizQuestionGroupSerializer
    permission_classes = [permissions.IsAdminUser]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["quiz"]


@api_view(["GET"])
def questionIds(request):
    quiz_id = request.GET.get("quiz_id")

    queryset = QuizQuestion.objects.filter(quiz_id=quiz_id)
    question_ids = list(queryset.values_list("question_id", flat=True))

    max_order_number = queryset.aggregate(max_order=Max("order_number"))["max_order"]
    max_order_number = max_order_number + 1 if max_order_number is not None else 1

    data = {"numbers": question_ids, "max_order_number": max_order_number}
    return Response(data)


@api_view(["PUT"])
def swapQuestions(request):
    id1 = request.data.get("object1")
    id2 = request.data.get("object2")
    obj1 = QuizQuestion.objects.get(id=id1)
    obj2 = QuizQuestion.objects.get(id=id2)
    temp = obj1.order_number
    obj1.order_number = obj2.order_number
    obj2.order_number = temp
    obj1.save()
    obj2.save()
    return Response(status=200)


@api_view(["PUT"])
def swapQuestionGroups(request):
    id1 = request.data.get("object1")
    id2 = request.data.get("object2")
    obj1 = QuizQuestionGroup.objects.get(id=id1)
    obj2 = QuizQuestionGroup.objects.get(id=id2)
    temp = obj1.order_number
    obj1.order_number = obj2.order_number
    obj2.order_number = temp
    obj1.save()
    obj2.save()
    return Response(status=200)


class QuizQuestionBulkView(
    generics.ListCreateAPIView, generics.DestroyAPIView, generics.UpdateAPIView
):
    queryset = QuizQuestion.objects.all()
    serializer_class = BulkQuizQuestionSerializer
    permission_classes = [permissions.IsAdminUser]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data, many=True)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(status=201, headers=headers)

    def update(self, request, *args, **kwargs):
        object_ids = request.data.pop("objects")
        queryset = self.get_queryset().filter(id__in=object_ids)
        queryset.update(**request.data)

        return Response(status=200)

    def destroy(self, request, *args, **kwargs):
        question_ids = [question["id"] for question in request.data]
        queryset = self.get_queryset().filter(question_id__in=question_ids)
        self.perform_destroy(queryset)
        return Response(status=204)


class AllowedGradeApiView(APIView):
    def get(self, request, pk):
        data = AllowedGrade.objects.filter(quiz__id=pk).values_list(
            "grade__id", flat=True
        )
        return Response({"values": data})

    def post(self, request, pk):
        ids = request.data.get("values")
        quiz = Quiz.objects.get(pk=pk)
        allowed_grade_list = [AllowedGrade(quiz=quiz, grade_id=id) for id in ids]
        AllowedGrade.objects.bulk_create(allowed_grade_list)

        excludeUsers = AllowedUser.objects.filter(quiz=quiz).values_list(
            "user_id", flat=True
        )
        allowed_user_list = []
        for id in ids:
            users = User.objects.filter(grade_id=id).exclude(id__in=excludeUsers)
            allowed_user_list.extend(
                AllowedUser(quiz=quiz, user=user) for user in users
            )
        AllowedUser.objects.bulk_create(allowed_user_list)
        return Response({"status": "created"})

    def delete(self, request, pk):
        ids = request.data.get("values")
        users = AllowedUser.objects.filter(quiz_id=pk, user__grade_id__in=ids)
        users.delete()

        allowed_grades = AllowedGrade.objects.filter(quiz_id=pk, grade_id__in=ids)
        allowed_grades.delete()

        return Response({"status": "deleted"})


class AllowedUserViewSet(viewsets.ModelViewSet):
    queryset = AllowedUser.objects.all()
    serializer_class = AllowedUserSerializer
    permission_classes = [permissions.IsAdminUser]
    pagination_class = QuizPagination
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["quiz", "user__grade_id"]


class AllowedUserBulkView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request, pk):
        quiz = Quiz.objects.get(pk=pk)
        ids = AllowedUser.objects.filter(quiz=quiz).values_list("user_id", flat=True)
        return Response({"values": ids})

    def post(self, request, pk):
        quiz = Quiz.objects.get(pk=pk)
        ids = request.data.get("values")
        allowed_user_list = [
            AllowedUser(quiz=quiz, user=user)
            for user in User.objects.filter(id__in=ids)
        ]
        AllowedUser.objects.bulk_create(allowed_user_list)
        return Response({"status": "created"})

    def delete(self, request, pk):
        quiz = Quiz.objects.get(pk=pk)
        ids = request.data.get("values")
        allowed_users = AllowedUser.objects.filter(quiz=quiz, user_id__in=ids)
        allowed_users.delete()
        return Response({"status": "deleted"})


class UserQuizView(generics.ListAPIView):
    serializer_class = UserQuizSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = Quiz.objects.filter(
            Q(allowed_users__user_id=user.id, access="private") | Q(access="public")
        ).distinct()
        return queryset

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context["request"] = self.request
        return context

    def get(self, request, *args, **kwargs):
        pk = kwargs.get("pk")

        if pk:
            obj = get_object_or_404(self.get_queryset(), pk=pk)
            serializer = self.get_serializer(obj)
            return Response(serializer.data)

        return super().get(request, *args, **kwargs)


class StartQuizView(APIView):
    permission_classes = [permissions.IsAuthenticated, StartQuizPermission]

    def get_object(self, pk):
        try:
            return Quiz.objects.get(pk=pk)
        except Quiz.DoesNotExist:
            return Http404

    def get(self, request, pk):
        quiz = self.get_object(pk)
        user = request.user

        attempt = create_user_attempt(user, quiz)
        if attempt:
            serializer = UserAttemptSerializer(
                attempt, context={"request": self.request}
            )
            return Response(serializer.data)
        else:
            return Response({"status": "error"}, status=500)
